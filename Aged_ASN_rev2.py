# Filename: Aged_ASN_rev2
# Author: Sophia Hammerbeck (Sophia.Hammerbeck@geappliances)
# Credits: Pieced together from code by Zachary Ramirez
#          (Zachary.Ramirez@geappliances.com)
# Date: 3/2/2023
#
# Purpose: Evaluates Firm Order Reports to prioritize based on age of ASN
#          and generates summary report.
# Comments: See Weeks on Hand Tool for original code.


import openpyxl as xl
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
import datetime as dt
from openpyxl.styles import fills


def initialize():
    """Gather Firm Order Report from user input."""
    fid = "1"
    while fid == "1":
        print("\n")
        fid = str(input("Enter the file path for the Firm Order Report "
                        "(Press 1 for help or 0 to exit): "))
        if fid == "1":
            print("\nTo find the file path:\n1) Open 'Files'\n2) Right Click on "
                  "the Weeks on Hand Report\n3) Select 'Copy as path'\n4) Paste "
                  "the path into the terminal when prompted")
        elif fid[0] == '"':
            fid = fid[1:-1]
    #fid = r"C:\git\mat\RandomAutomationTools\AP3_Materials\test dummy db6.xlsx"
    return (fid)


def valid(fid):
    """Validate file input and print error message if needed."""
    validate = 0
    if fid == "0":
        print("Closing the program.")
    elif len(fid) < 15:
        print("Error 405: Invalid file path.")
    elif fid[-3:] == "xls":
        print(
            "Error 603: This program does not support the old .xls file format. "
            "Please convert it to the more recent .xlsx file format.")
    elif fid[0:5] == "https":
        print("Error 604: This program cannot access files stored online. "
              "Please download the file to your computer. ")
    else:
        validate = 1
    return (validate)


def in_scope(fid):
    """Retrieve a list of open POs."""

    print("Defining scope...", end='', flush=True)
    PO_matrix_1 = pd.read_excel(fid)
    if PO_matrix_1.columns[0].lower() != "po number":
        new_header = PO_matrix_1.iloc[0]
        PO_matrix_1 = PO_matrix_1[1:]
        PO_matrix_1.columns = new_header

    listx = ['PARTIALLY RECEIVED', 'EXPECTED']
    PO_matrix_2 = PO_matrix_1[PO_matrix_1['ASN Status'].isin(listx)]
    wip_POs = PO_matrix_2.index.to_list()
    #Collect WIP POs, excluding cancelled POs.

    listy = ['FULLY RECEIVED']
    PO_matrix_3 = PO_matrix_1[PO_matrix_1['ASN Status'].isin(listy)]
    full_POs = PO_matrix_3.index.to_list()
    #Collect fully received POs (Must be separate because ALL fully received POs
    #get coded green.

    print("Done")
    return (wip_POs, full_POs)


def get_headers(ws):
    """Retrieve the relevant headers."""
    print("Retrieving headers...", end='', flush=True)
    header_cols = ['PO Number', 'Vendor Name', 'Due Date']
    counter = 0
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            if "PO Number" == ws.cell(row, col).value:
                number_col = col
                counter += 1
            elif "Vendor Name" == ws.cell(row, col).value:
                supplier_col = col
                counter += 1
            elif "Estimated Receipt Date (ETA)" == ws.cell(row, col).value:
                due_date_col = col
                counter += 1
        if counter == 3:
            break
    print("Done")
    return (row, number_col, supplier_col, due_date_col, header_cols)


def evaluate(ws, POs, full_POs, header_row, number_col,
             supplier_col, due_date_col, header_cols):
    """Color code cells based on the days since the ASN due date."""
    print("Evaluating Firm Order Report...", end='', flush=True)
    for row in range(header_row + 1, ws.max_row + 1):
        if (ws.cell(row, 1) == None):
            ws.cell(row, 1).value = ws.cell(row - 1, 1).value
            # If a cell is empty, fill it with the value directly above.
        date = ws.cell(row, due_date_col).value
        PO_row = row - 2
        if date is not None:
            if PO_row in POs:
                date = dt.datetime.strptime(str(date), '%Y-%m-%d %H:%M:%S').date()
                days_since = date - dt.date.today()
                days_since = int(days_since.days)
                for col, col_name in enumerate(header_cols):
                    col = col + 1  # Need col to be 1-indexed instead of 0-indexed
                    if days_since <= -30:
                        ws.cell(row, col).fill = fills.PatternFill('solid',
                                                                   fgColor='FF0000')
                    elif days_since >= 0:
                        ws.cell(row, col).fill = PatternFill(start_color='00FF00',
                                                             end_color='00FF00',
                                                             fill_type='solid')
                    else:
                        ws.cell(row, col).fill = PatternFill(start_color='FFFF00',
                                                             end_color='FFFF00',
                                                             fill_type='solid')
            elif PO_row in full_POs:
                for col, col_name in enumerate(header_cols):
                    col = col + 1
                    ws.cell(row, col).fill = PatternFill(start_color='00FF00',
                                                         end_color='00FF00',
                                                         fill_type='solid')
                    #All full_POs must be green.

   # POs.remove(ws.cell(row,1).value)
    print("Done")
    return ()


def make_table(ws, header_row, name):
    """Create & stylize table."""
    print("Formatting Results...", end='', flush=True)
    if ws.tables == {}:
        alphabet = " ABCDEFGHIJKLMNOPQRSTUVWXYZ"
        start = f"A{header_row}:"

        if (ws.max_column + 1) < 26:
            end = f"{alphabet[(ws.max_column)]}{(ws.max_row)}"
        else:
            end = f"{alphabet[(int((ws.max_column) / 26))]}" \
                  f"{alphabet[((ws.max_column) % 26)]}{(ws.max_row)}"
        ran = start + end
        tab = Table(displayName=f"Table_{name}", ref=ran)
        style = TableStyleInfo(name="TableStyleMedium9",
                               showFirstColumn=False, showLastColumn=False,
                               showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        return (tab)
    else:
        return (0)


def get_summary(ws, header_row, supplier_col, header_cols):
    """For each supplier, collect a summary of how many Red/Yellow/Green cells
       it has."""
    print("Retrieving Summary Information...", end='', flush=True)
    summary_info = {}
    for row in range(header_row + 1, ws.max_row + 1):
        if ws.cell(row, supplier_col).value is not None \
                and ws.cell(row, supplier_col).value != "0":
            if ws.cell(row, supplier_col).value not in summary_info.keys():
                summary_info[ws.cell(row, supplier_col).value] = {"Red": 0,
                                                                  "Yellow": 0,
                                                                  "Green": 0}
                # Generates nested dictionaries containing initialized
                # information for each supplier
        if ws.cell(row, 1).fill.fgColor.rgb == "00FF0000":
            summary_info[ws.cell(row, supplier_col).value]["Red"] += 1
            # Updates the red dictionary for each supplier and each site
        elif ws.cell(row, 1).fill.fgColor.rgb == "00FFFF00":
            summary_info[ws.cell(row, supplier_col).value]["Yellow"] += 1
            # Updates the yellow dictionary for each supplier and each site
        elif ws.cell(row, 1).fill.fgColor.rgb == "0000FF00":
            summary_info[ws.cell(row, supplier_col).value]["Green"] += 1
            # Updates the green dictionary for each supplier and each site
    print("Done")
    return (summary_info)


def make_summary(wb, summary_info, fid):
    """Create a table displaying the data collected in get_summary."""
    headers = ["Supplier", "Red Qty", "Yellow Qty", "Green Qty"]
    print("Generating Summary...", end='', flush=True)

    # 1: Creates summary sheet and removes outdated information
    if 'Summary' in wb.sheetnames:
        wb.remove(wb['Summary'])
    wb.create_sheet('Summary')
    ws = wb['Summary']
    wb.save(filename=fid)

    # 2: Formats headers
    for index in range(0, len(headers)):
        ws.cell(1, index + 1).value = headers[index]
        ws.cell(1, index + 1).fill.fgColor.rgb = "00FFFFCC"

    # 3: Enters summary info
    MAX_ROW = 1
    row = 2
    for row, supplier in enumerate(summary_info.keys()):
        row += 2
        # print(row)
        ws.cell(row, 1).value = supplier
        if ws.cell(row, 1).value is None:
            ws.cell(row, 1).value = ws.cell(row - 1, 1).value
        ws.cell(row, 2).value = summary_info[supplier]['Red']
        ws.cell(row, 3).value = summary_info[supplier]['Yellow']
        ws.cell(row, 4).value = summary_info[supplier]['Green']
    print("Done")
    tab = make_table(ws, 1, 2)
    if tab != 0:
        ws.add_table(tab)
        wb.save(filename=fid)
    print("Done")
    return ()


def main():
    """Execute the program and print errors if needed."""
    fid = initialize()
    cont = valid(fid)
    # Checks exit criterion
    while cont == 1:
        try:
            wb = xl.load_workbook(filename=fid)
            ws = wb.active
            POs, full_POs = in_scope(fid)
            (header_row, number_col, supplier_col, due_date_col, header_cols) = get_headers(ws)
            tab = make_table(ws, header_row, 1)
            if tab != 0:
                ws.add_table(tab)
            wb.save(filename=fid)
            wb.close
            wb = xl.load_workbook(filename=fid)
            ws = wb.active
            evaluate(ws, POs, full_POs, header_row, number_col, supplier_col, due_date_col, header_cols)
            summary_info = get_summary(ws, header_row, supplier_col, header_cols)
            make_summary(wb, summary_info, fid)
            print("Done")
            print("Done\nIt is now safe to open the Firm Order Report.")
            print("\n***************************************************************\n")
            print("Table Key:")
            print("Red Cell: Over a month overdue")
            print("Yellow Cell: Less than a month overdue")
            print("Green Cell: Not yet due OR fully received")
            print("\n***************************************************************")
        except xl.utils.exceptions.InvalidFileException:
            print("Error 403: This file is not present on the device.")
        except PermissionError:
            print("Error 303: Please close the file and try again. Changes have not been saved.")
        cont = 0
        wb.close()
    return ()


if __name__ == "__main__":
    main()
    input("\nPress 'Enter' to safely close the program. ")